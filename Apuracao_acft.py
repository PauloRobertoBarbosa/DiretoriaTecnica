import openpyxl as op
import pandas as pd
from pathlib import Path
import os
os.chdir(r"C:\Users\paulo.roberto\Documents\Copied")
path = Path(r"R:\Corporativo\Diretoria Tecnica\1-Gestão de Custos\Diretoria de Manutenção\Custo de Manutenção\ACTF-WO\Teste Apuração")
lista = []
for top, dirs, files in os.walk(path):
    for xls in files:
        if xls.lower().__contains__("xls"):
            lista.append(os.path.join(top, xls))
materiais = ''
for mat in lista:
    if mat.lower().__contains__("materiais"):
        materiais = mat
hhexec = ''
for serv in lista:
    if serv.lower().__contains__("hh"):
        hhexec = serv

pdmat = pd.ExcelFile(materiais)
dfmat = pdmat.parse()
pdpac = pd.ExcelFile(r"R:\Corporativo\Diretoria Tecnica\1-Gestão de Custos\Realizado\Actual-2019\6.Jun-19\Relatório de preço de itens - 06-2019.xlsx")
pac = pdpac.parse()
pdserv = pd.ExcelFile(hhexec)
dfserv = pdserv.parse()

dicpn = {}
dicwo = {}
contpac = 0
contservdic = 0

for i in pac["PN"]:
    dicpn.update({pac["PN"].values[contpac]: [pac["CUSTO_PAC_ITEM"].values[contpac], pac["PRECO_UNITARIO_ITEM"].values[contpac], pac["PRECO_ULTIMA_COMPRA"].values[contpac], pac["CATEGORIA_TECNICA"].values[contpac]]})
    contpac += 1

for i in dfserv["wo"]:
    dicwo.update({dfserv["wo"].values[contservdic]: [dfserv["wo_description"].values[contservdic]]})
    contservdic += 1

arquivo = r"C:\Users\paulo.roberto\Documents\Apuração.xlsx"
planworkscope = op.load_workbook(arquivo)
ws = planworkscope.active
maiorrow= ws.max_row+1
cont = 0
pnatual = ''
valormo = 78.20
valormolessor = 64
valordolar = 3.881
# print(dfmat.head(0))
for i in dfmat["TASK_CARD"]:
    ws["A{}".format(maiorrow)] = dfmat["AC"].values[cont]
    ws["B{}".format(maiorrow)] = dfmat["WO"].values[cont]
    if dfmat["WO"].values[cont] in dicwo.keys():
        ws["C{}".format(maiorrow)] = dicwo[dfmat["WO"].values[cont]][0]
    else:
        ws["C{}".format(maiorrow)] = 'NA'
    ws["D{}".format(maiorrow)] = dfmat["TASK_CARD"].values[cont]
    ws["E{}".format(maiorrow)] = 'NA'
    ws["F{}".format(maiorrow)] = dfmat["TASK_CARD_DESCRIPTION"].values[cont]
    ws["G{}".format(maiorrow)] = dfmat["PN"].values[cont]
    ws["H{}".format(maiorrow)] = dfmat["PN_DESCRIPTION"].values[cont]
    ws["I{}".format(maiorrow)] = dfmat["QTY"].values[cont]
    ws["J{}".format(maiorrow)] = dicpn['{}'.format(dfmat["PN"].values[cont])][3]

    pnatual = dfmat["PN"].values[cont]
    regra1 = []
    if dfmat["PN"].values[cont] in dicpn.keys():
        if dicpn['{}'.format(dfmat["PN"].values[cont])][0] == 0 and dicpn['{}'.format(dfmat["PN"].values[cont])][1] == 0:
            ws["K{}".format(maiorrow)] = dicpn[pnatual][2]
            ws["L{}".format(maiorrow)] = dicpn[pnatual][2] * dfmat["QTY"].values[cont]
            ws["S{}".format(maiorrow)] = dicpn[pnatual][2] * dfmat["QTY"].values[cont]
            ws["T{}".format(maiorrow)] = dicpn[pnatual][2] * dfmat["QTY"].values[cont]
            ws["U{}".format(maiorrow)] = dicpn[pnatual][2] * dfmat["QTY"].values[cont] / valordolar
            ws["V{}".format(maiorrow)] = dicpn[pnatual][2] * dfmat["QTY"].values[cont] / valordolar
        if dicpn['{}'.format(dfmat["PN"].values[cont])][0] == 0 and dicpn['{}'.format(dfmat["PN"].values[cont])][2] == 0:
                ws["K{}".format(maiorrow)] = dicpn[pnatual][1]
                ws["L{}".format(maiorrow)] = dicpn[pnatual][1] * dfmat["QTY"].values[cont]
                ws["S{}".format(maiorrow)] = dicpn[pnatual][1] * dfmat["QTY"].values[cont]
                ws["T{}".format(maiorrow)] = dicpn[pnatual][1] * dfmat["QTY"].values[cont]
                ws["U{}".format(maiorrow)] = dicpn[pnatual][1] * dfmat["QTY"].values[cont] / valordolar
                ws["V{}".format(maiorrow)] = dicpn[pnatual][1] * dfmat["QTY"].values[cont] / valordolar
        if dicpn['{}'.format(dfmat["PN"].values[cont])][0] == 0 and dicpn['{}'.format(dfmat["PN"].values[cont])][1] != 0:
                ws["K{}".format(maiorrow)] = dicpn[pnatual][1]
                ws["L{}".format(maiorrow)] = dicpn[pnatual][1] * dfmat["QTY"].values[cont]
                ws["S{}".format(maiorrow)] = dicpn[pnatual][1] * dfmat["QTY"].values[cont]
                ws["T{}".format(maiorrow)] = dicpn[pnatual][1] * dfmat["QTY"].values[cont]
                ws["U{}".format(maiorrow)] = dicpn[pnatual][1] * dfmat["QTY"].values[cont] / valordolar
                ws["V{}".format(maiorrow)] = dicpn[pnatual][1] * dfmat["QTY"].values[cont] / valordolar
        if dicpn['{}'.format(dfmat["PN"].values[cont])][0] != 00:
                ws["K{}".format(maiorrow)] = dicpn[pnatual][0]
                ws["L{}".format(maiorrow)] = dicpn[pnatual][0] * dfmat["QTY"].values[cont]
                ws["S{}".format(maiorrow)] = dicpn[pnatual][0] * dfmat["QTY"].values[cont]
                ws["T{}".format(maiorrow)] = dicpn[pnatual][0] * dfmat["QTY"].values[cont]
                ws["U{}".format(maiorrow)] = dicpn[pnatual][0] * dfmat["QTY"].values[cont] / valordolar
                ws["V{}".format(maiorrow)] = dicpn[pnatual][0] * dfmat["QTY"].values[cont] / valordolar

        if str(dfmat["TASK_CARD"].values[cont])[0:2] == str('NR'):
            ws["P{}".format(maiorrow)] = 'Non Routine'
        else:
            ws["P{}".format(maiorrow)] = 'Routine'

        ws["Q{}".format(maiorrow)] = 'Material'

        cont += 1
        maiorrow += 1
        pnatual = ''

contserv = 0
maiorrowserv = ws.max_row

for x in dfserv["task_card"]:
    ws["A{}".format(maiorrowserv)] = dfserv["ac"].values[contserv]
    ws["B{}".format(maiorrowserv)] = dfserv["wo"].values[contserv]
    ws["C{}".format(maiorrowserv)] = dfserv["wo_description"].values[contserv]
    ws["D{}".format(maiorrowserv)] = dfserv["task_card"].values[contserv]
    ws["E{}".format(maiorrowserv)] = dfserv["schedule_task_card"].values[contserv]
    ws["F{}".format(maiorrowserv)] = dfserv["task_card_description"].values[contserv]
    ws["M{}".format(maiorrowserv)] = dfserv["h_h_executado"].values[contserv]
    ws["N{}".format(maiorrowserv)] = float(dfserv["h_h_executado"].values[contserv]) * float(valormo)
    ws["S{}".format(maiorrowserv)] = float(dfserv["h_h_executado"].values[contserv]) * float(valormo)
    ws["O{}".format(maiorrowserv)] = float(dfserv["h_h_executado"].values[contserv]) * float(valormo) / float(valordolar)
    ws["U{}".format(maiorrowserv)] = float(dfserv["h_h_executado"].values[contserv]) * float(valormo) / float(valordolar)

    if str(dfserv["task_card"].values[contserv])[0:2] == str('NR'):
        ws["P{}".format(maiorrowserv)] = 'Non Routine'
    else:
        ws["P{}".format(maiorrowserv)] = 'Routine'

    ws["Q{}".format(maiorrowserv)] = 'Labor'
    ws["R{}".format(maiorrowserv)] = float(dfserv["h_h_executado"].values[contserv]) * float(valormolessor) * float(valordolar)
    ws["T{}".format(maiorrowserv)] = float(dfserv["h_h_executado"].values[contserv]) * float(valormolessor) * float(valordolar)
    ws["V{}".format(maiorrowserv)] = float(dfserv["h_h_executado"].values[contserv]) * float(valormolessor)

    contserv +=1
    maiorrowserv +=1

planworkscope.save(arquivo)