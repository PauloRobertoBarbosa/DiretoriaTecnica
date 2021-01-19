import openpyxl as xly
import os
os.chdir(r'C:\Users\paulo.roberto\Documents\Compilado PBI')
path = os.listdir()
fileopen = []
for i in path:
    if i.__contains__('~$'):
        pass
    lista_sheets = []
    plan_atual = xly.load_workbook(i)
    for x in plan_atual.sheetnames:
        lista_sheets.append(x)
    for sheet in lista_sheets:
        if sheet == "Budget2020":
            pass
        else:
            value = sheet[0:4]
            plan = xly.load_workbook(i)
            ss_sheet = plan[value]
            ss_sheet.title = 'Budget2020'
            plan.save(i)
            plan.close()