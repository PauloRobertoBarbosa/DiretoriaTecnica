import xlrd
import openpyxl as op
from pathlib import Path
import os
import shutil

os.chdir(r"C:\Users\paulo.roberto\OneDrive - Azul Linhas Aéreas\Workscope")
path = Path(r"C:\Users\paulo.roberto\OneDrive - Azul Linhas Aéreas\Workscope")
lista = []
for top, dirs, files in os.walk(path):
    for xls in files:
        if xls.__contains__("~$")==False and xls.lower().__contains__("xls") and xls.lower().__contains__("workscope"):
            lista.append(os.path.join(top, xls))
            # print(os.path.join(top, xls))
            # print(os.path.getmtime(xls))


newpath = Path(r"C:\Users\paulo.roberto\Documents\Copied 2018")
count = 0
for item in lista:
    shutil.copy(lista[count], newpath)
    count +=1


# print(*lista, sep="\n")

#
# listausar = list(range(0, len(lista)))
#
# for sheetusar in listausar:
#     try:
#         sheets = str(xlrd.open_workbook(lista[sheetusar], encoding_override='cp1252').sheet_names())
#         workscope = ''
#     # print(sheetusar)
#         for sheet in sheets:
#             if str(sheet.lower())=='workscope':
#                 workscope = str(sheet)
#             # print(sheet)
#                 filez = xlrd.open_workbook(lista[sheetusar], encoding_override='cp1252')
#     except:
#         pass
#     try:
#         book = filez.sheet_by_name(workscope)
#         col1 = []
#         wo = []
#         wofim = []
#         ncol = 0
#         while book.ncols > ncol:
#             for val in book.col(ncol):
#                 if val.value != '':
#                     col1.append(val.value)
#             ncol +=1
#         acft = col1[col1.index('TAIL NUMBER')+1]
#         acftmodel = col1[col1.index('ACFT MODEL')+1]
#         tipo = ''
#         for t in col1:
#             if str(t).__contains__("- MAJOR")==True:
#                 tipo = t
#         checktype = col1[col1.index(str(tipo))+1]
#         mro = col1[col1.index('MRO')+1]
#         plantat = col1[col1.index('PLANNED TAT')+1]
#         realtat = col1[col1.index('REAL TAT')+1]
#         schstartdate = col1[col1.index('SCHEDULE START DATE') + 1]
#         schcompletion = col1[col1.index('SCHEDULE COMPLETION') + 1]
#         realcompletion = col1[col1.index('REAL COMPLETION DATE') + 1]
#
#         for numwo in col1:
#             if type(numwo)==float:
#                 if len(str(numwo))==8:
#                     wo.append(numwo)
#
#         for iwo in wo:
#             if iwo in wofim:
#                 pass
#             else:
#                 wofim.append(iwo)
#         arquivo = r"C:\Users\paulo.roberto\Documents\worscopeconsolidado 2018.xlsx"
#         planworkscope = op.load_workbook(arquivo)
#         ws = planworkscope["woworkscope"]
#         maiorrow= ws.max_row+1
#
#         for valcel in wofim:
#             ws['A{}'.format(maiorrow)] = acft
#             ws['B{}'.format(maiorrow)] = acftmodel
#             ws['C{}'.format(maiorrow)] = checktype
#             ws['D{}'.format(maiorrow)] = mro
#             ws['E{}'.format(maiorrow)] = plantat
#             ws['F{}'.format(maiorrow)] = realtat
#             ws['G{}'.format(maiorrow)] = valcel
#             ws['H{}'.format(maiorrow)] = schstartdate
#             ws['I{}'.format(maiorrow)] = schcompletion
#             ws['J{}'.format(maiorrow)] = realcompletion
#             ws['K{}'.format(maiorrow)] = lista[sheetusar]
#             maiorrow += 1
#
#         planworkscope.save(arquivo)
#
#     except:
#         pass