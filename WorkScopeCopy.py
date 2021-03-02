#######################COPIA WORKSCOPES DA REDE ####################
import xlrd
import openpyxl as op
from pathlib import Path
import os
import pandas as pd

path = Path(r"C:\Users\paulo.roberto\OneDrive - Azul Linhas Aéreas\Workscope\Workscope 2021")
lista = []
listaxls = []
lista2 = []
dftotal = pd.DataFrame(columns=["AC",	"MODEL",	"CHECKTYPE",	"MRO",	"TAT PLAN",	"TAT REAL",	"WO",	"EC",	"TASK DESCRIPTION",	"SCHEDULE START DATE",	"SCHEDULE COMPLETION DATE",	"REAL COMPLETION",	"PATH"])

for top, dirs, files in os.walk(path):
    for xls in files:
        if xls.__contains__("~$")==False and xls.__contains__(".pdf")==False and xls.lower().__contains__("workscope") and xls.lower().__contains__("xls"):
            lista.append(os.path.join(top, xls))

listausar = list(range(0, len(lista)))

listautilizada = []
for sheetusar in listausar:
    sheetname = xlrd.open_workbook(lista[sheetusar])
    sheets = sheetname.sheet_names()

    sheetabrir = 'WorkScope'
    for sheet in sheets:
        if sheet.lower()=='workscope':
            filez = xlrd.open_workbook(lista[sheetusar], encoding_override='cp1252')
            listautilizada.append(sheet)
            sheetabrir = sheet

    book = filez.sheet_by_name(sheetabrir)
    col1 = []
    wo = []
    ec = []
    taskdesk = []
    wofim = []
    ncol = 0
    while book.ncols > ncol:
        for val in book.col(ncol):
            if val.value != '':
                col1.append(val.value)
        ncol +=1

    acft = col1[col1.index('TAIL NUMBER')+1]
    acftmodel = col1[col1.index('ACFT MODEL')+1]
    tipo = ''
    for t in col1:
        if str(t).__contains__("- MAJOR")==True:
            tipo = t
    checktype = col1[col1.index(str(tipo))+1]
    mro = col1[col1.index('MRO')+1]
    plantat = col1[col1.index('PLANNED TAT')+1]
    realtat = col1[col1.index('REAL TAT')+1]
    schstartdate = col1[col1.index('SCHEDULE START DATE') + 1]
    schcompletion = col1[col1.index('SCHEDULE COMPLETION') + 1]
    realcompletion = col1[col1.index('REAL COMPLETION DATE') + 1]

    contwo = 0
    contindex = 0
    for cont in book.col(1):
        contindex +=1
        if str(cont) == "number:2.0":
            contwo = contindex

    while contwo < len(book.col(1)):
        if len(str(book.col(1)[contwo])) == 15:
            wo.append(str(book.col(1)[contwo]))
            ec.append(str(book.col(7)[contwo]))
            taskdesk.append(str(book.col(15)[contwo]))
        contwo += 1

    # arquivo = r"C:\Users\paulo.roberto\Documents\worscopeconsolidado.xlsx"
    # planworkscope = op.load_workbook(arquivo)
    # ws = planworkscope["woworkscope"]
    # maiorrow= ws.max_row+1
    # contec = 0

    # for valcel in wo:
    #     ws['A{}'.format(maiorrow)] = acft
    #     ws['B{}'.format(maiorrow)] = acftmodel
    #     ws['C{}'.format(maiorrow)] = checktype
    #     ws['D{}'.format(maiorrow)] = mro
    #     ws['E{}'.format(maiorrow)] = plantat
    #     ws['F{}'.format(maiorrow)] = realtat
    #     ws['G{}'.format(maiorrow)] = valcel
    #     ws['H{}'.format(maiorrow)] = ec[contec]
    #     ws['I{}'.format(maiorrow)] = taskdesk[contec]
    #     ws['J{}'.format(maiorrow)] = schstartdate
    #     ws['K{}'.format(maiorrow)] = schcompletion
    #     ws['L{}'.format(maiorrow)] = realcompletion
    #     ws['M{}'.format(maiorrow)] = lista[sheetusar]
    #     maiorrow += 1
    #     contec += 1
    #
    # planworkscope.save(arquivo)

    columns = ["WO", "EC", "TASK DESCRIPTION"]

    df = pd.DataFrame(zip(wo,ec, taskdesk),columns=columns)
    df["AC"] = acft
    df["MODEL"] = acftmodel
    df["CHECKTYPE"] = checktype
    df["MRO"] = mro
    df["TAT PLAN"] = plantat
    df["TAT REAL"] = realtat
    df["SCHEDULE START DATE"] = schstartdate
    df["SCHEDULE COMPLETION DATE"] = schcompletion
    df["REAL COMPLETION"] = realcompletion
    df["PATH"]= lista[sheetusar]

    dftotal = dftotal.append(df)

excel = pd.ExcelWriter(r'C:\Users\paulo.roberto\Documents\worscopeconsolidado2.xlsx',engine='xlsxwriter')
dftotal.to_excel(excel, sheet_name='Consolidado', index=False)
excel.save()